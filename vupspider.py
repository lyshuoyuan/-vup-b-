#首先cookie需要填入cookie（也可以尝试一下不填能不能用，咱没试，用了一个小号的cookie）
#然后是结果会生成一个Excel表，需要去修改一下自己的地址给的是一个范例格式（已经成功跑出来了结果，但是不知道会不会没有文件夹的时候自动创建，所以还是修改一下好了）
#这个代码用来获取虚拟主播区所有主播的直播间名称，主播名称，房间id和主播id，本来希望通过直播间名称来分析直播内容，然后发现自己年轻了
#vuplist获取的某一个页面上的主播信息
#最下面的resultlist 提取和合并后的直播间名称，主播名称，房间id和主播id

import requests
import json
import openpyxl
import time

def getvuplist(num):       #num是页面数（第几页）
    url = "https://api.live.bilibili.com/room/v3/area/getRoomList?platform=web&parent_area_id=9&cate_id=0&area_id=0&sort_type=live_time&page="+str(num)+"&page_size=30&tag_version=1"
    headers = {
        "accept":"application/json, text/plain, */*",
        "accept-encoding":"gzip, deflate, br",
        "accept-language":"zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "cookie":"",                             #这里填cookie
        "origin":"https://live.bilibili.com",
        "referer":"https://live.bilibili.com/",
        "sec-fetch-dest":"empty",
        "sec-fetch-mode":"cors",
        "sec-fetch-site":"same-site",
        "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.183 Safari/537.36 Edg/86.0.622.63"
    }
    r = requests.get(url=url,headers = headers)
    data = json.loads(r.text)
    vuplist = data['data']['list']
    return vuplist

def getdatalist(vuplist):#上一步的结果返回的一个，json？list？还是什么的一个格式的东西
    resultlist=[]
    for i in range(len(vuplist)):
        vupdic=[]
        vupdic.append(vuplist[i]['uname'])
        vupdic.append(vuplist[i]['roomid'])
        vupdic.append(vuplist[i]['uid'])
        vupdic.append(vuplist[i]['title'])
        resultlist.append(vupdic)
    return resultlist

def saveasxlsx(resultlist,dt):#结果的写入，resultlist就是结果了，这步只是写入Excel，dt表示时间
    
    wb = openpyxl.Workbook()
    sheet1 = wb.create_sheet('虚拟主播分区信息爬取',0)
    title_ls = ['用户名','房间号','主播id','直播间名称'] 
    
    for i in range(len(title_ls)):
        sheet1.cell(1,1+i).value = title_ls[i]
        
    for row in range(len(resultlist)):  
        for column in range(len(resultlist[row])):
            sheet1.cell(row+2,column+1).value = resultlist[row][column]
            
        
    time1 = time.strftime("%Y-%m-%d %H_%M_%S",dt) 
    
    wb.save('F:/编程/ddscrip/虚拟主播直播区'+time1+'数据'+'.xlsx')         #这里修改文件路径和文件名
    wb.close()
    return True

resultlist = []
vuplist = getvuplist(0)
i=0
while len(vuplist)!=0:
    i+=1
    vuplist = getvuplist(i)
    resultlist+=getdatalist(vuplist)

allupnum = len(resultlist)
dt=time.localtime()    
saveasxlsx(resultlist,dt)